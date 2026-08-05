from __future__ import annotations

from pathlib import Path
from typing import Iterable
import csv
import io

from openpyxl import load_workbook

from .models import Card


ALIASES: dict[str, set[str]] = {
    "title": {"title", "name", "label", "item", "topic", "age", "card"},
    "value": {"value", "amount", "score", "number", "rank", "percentage", "percent"},
    "description": {"description", "desc", "details", "detail", "explanation", "subtitle", "text"},
    "image": {"image", "img", "picture", "photo", "icon", "image_url", "image path", "image_path", "url"},
}


def _text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "True" if value else "False"
    return str(value).strip()


def _normalize(value: object) -> str:
    return _text(value).lower().replace("-", "_")


def _map_headers(headers: Iterable[object]) -> dict[int, str]:
    mapped: dict[int, str] = {}
    used: set[str] = set()
    for index, header in enumerate(headers):
        normalized = _normalize(header)
        for target, aliases in ALIASES.items():
            if target not in used and normalized in aliases:
                mapped[index] = target
                used.add(target)
                break
    return mapped


def _looks_like_header(row: list[object], header_map: dict[int, str]) -> bool:
    nonempty = [_normalize(cell) for cell in row if _text(cell)]
    if len(header_map) >= 2:
        return True
    if len(header_map) == 1 and len(nonempty) == 1:
        # A one-column spreadsheet headed only by "Title"/"Value" is valid,
        # but "Age,10" and "Card,1" are data and must not lose their first row.
        return nonempty[0] in {"title", "value", "description", "image", "image_url", "image_path"}
    return False


def _rows_to_cards(rows: list[list[object]], *, asset_base: Path | None = None) -> list[Card]:
    rows = [list(row) for row in rows if any(_text(cell) for cell in row)]
    if not rows:
        return []

    header_map = _map_headers(rows[0])
    has_named_headers = _looks_like_header(rows[0], header_map)
    data_rows = rows[1:] if has_named_headers else rows

    cards: list[Card] = []
    for row in data_rows:
        values = [_text(cell) for cell in row]
        if not any(values):
            continue
        if has_named_headers:
            mapped = {
                target: values[index] if index < len(values) else ""
                for index, target in header_map.items()
            }
        else:
            mapped = {
                "title": values[0] if len(values) > 0 else "",
                "value": values[1] if len(values) > 1 else "",
                "description": values[2] if len(values) > 2 else "",
                "image": values[3] if len(values) > 3 else "",
            }
        image = mapped.get("image", "")
        if image and asset_base is not None and not image.lower().startswith(("http://", "https://")):
            candidate = Path(image).expanduser()
            if not candidate.is_absolute():
                mapped["image"] = str((asset_base / candidate).resolve())
        cards.append(Card.from_mapping(mapped))
    return cards


def load_spreadsheet(path: str | Path) -> list[Card]:
    source = Path(path).expanduser().resolve()
    suffix = source.suffix.lower()
    if suffix == ".csv":
        with source.open("r", encoding="utf-8-sig", newline="") as handle:
            return _rows_to_cards(list(csv.reader(handle)), asset_base=source.parent)
    if suffix in {".xlsx", ".xlsm"}:
        workbook = load_workbook(source, read_only=True, data_only=True)
        try:
            sheet = workbook.active
            return _rows_to_cards(
                [list(row) for row in sheet.iter_rows(values_only=True)],
                asset_base=source.parent,
            )
        finally:
            workbook.close()
    raise ValueError("Cubical Create supports CSV, XLSX and XLSM files.")


def parse_pasted_data(text: str) -> list[Card]:
    cleaned = text.strip()
    if not cleaned:
        return []

    sample = cleaned[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",\t;|")
    except csv.Error:
        dialect = csv.excel_tab if "\t" in cleaned else csv.excel

    rows = list(csv.reader(io.StringIO(cleaned), dialect=dialect))
    return _rows_to_cards(rows)
