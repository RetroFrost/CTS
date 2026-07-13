from __future__ import annotations

import csv
import io
import json
import math
import re
import tempfile
from dataclasses import asdict, dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Iterable, Sequence


REFERENCE_REVEAL_SECONDS = 2.0
REFERENCE_SCROLL_SECONDS = 10.0 / 3.0
REFERENCE_END_HOLD_SECONDS = 2.0
REFERENCE_FADE_SECONDS = 0.8

MODEL_REFERENCE = "reference_detail"
MODEL_ILLUSTRATED = "illustrated_cards"
MODEL_CLASSIC = "classic_compact"
MODEL_DEFAULT_VISIBLE = {
    MODEL_REFERENCE: 4,
    MODEL_ILLUSTRATED: 3,
    MODEL_CLASSIC: 4,
}
VISIBLE_CARDS = 4  # Backwards-compatible reference-model constant.

# Each visual model owns the fields its layout actually renders. Values inside these
# columns may be blank; the schema exists to make the editor understandable.
MODEL_SCHEMAS: dict[str, tuple[tuple[str, str], ...]] = {
    MODEL_REFERENCE: (
        ("Badge Date / Value", "badge_primary"),
        ("Title", "title"),
        ("Description", "description"),
        ("Image", "image"),
    ),
    MODEL_ILLUSTRATED: (
        ("Badge Value", "badge_primary"),
        ("Badge Label", "badge_secondary"),
        ("Title", "title"),
        ("Artwork", "image"),
    ),
    MODEL_CLASSIC: (
        ("Value", "badge_primary"),
        ("Unit", "badge_secondary"),
        ("Title", "title"),
        ("Image", "image"),
    ),
}

FIELD_ROLES = ("badge_primary", "badge_secondary", "title", "description", "image")


class FriendlyError(RuntimeError):
    """An application error that is safe and useful to show to a user."""

    def __init__(self, summary: str, suggestion: str = "", details: str = ""):
        super().__init__(summary)
        self.summary = summary
        self.suggestion = suggestion
        self.details = details


@dataclass(slots=True)
class CardData:
    """Resolved renderer fields. Every field is optional by design."""

    uploaded: str = ""  # Primary badge value (historic field name kept for projects/API).
    title: str = ""
    description: str = ""
    image: str = ""
    badge_label: str = ""

    def is_blank(self) -> bool:
        return not any(
            value.strip()
            for value in (
                self.uploaded,
                self.title,
                self.description,
                self.image,
                self.badge_label,
            )
        )


@dataclass(slots=True)
class SpreadsheetData:
    headers: list[str] = field(default_factory=list)
    rows: list[list[str]] = field(default_factory=list)

    def normalized(self) -> "SpreadsheetData":
        width = max(len(self.headers), max((len(row) for row in self.rows), default=0))
        headers = _unique_headers(self.headers, width)
        rows = [
            [display_cell(row[index]) if index < len(row) else "" for index in range(width)]
            for row in self.rows
        ]
        return SpreadsheetData(headers, rows)

    def nonblank_rows(self) -> list[list[str]]:
        return [row for row in self.normalized().rows if any(value.strip() for value in row)]

    @property
    def row_count(self) -> int:
        return len(self.normalized().rows)


@dataclass(slots=True)
class AudioTrack:
    path: str = ""
    start_time: float = 0.0
    trim_start: float = 0.0
    trim_end: float | None = None
    volume: float = 1.0
    fade_in: float = 0.0
    fade_out: float = 0.0
    loop: bool = False

    def validate(self) -> None:
        if not self.path.strip():
            raise FriendlyError("A soundtrack row has no audio file.", "Choose a soundtrack file or remove the row.")
        if not Path(self.path).expanduser().is_file():
            raise FriendlyError(
                f"Soundtrack file not found: {self.path}",
                "Choose an existing audio file, then export again.",
            )
        values = (self.start_time, self.trim_start, self.volume, self.fade_in, self.fade_out)
        if any(not math.isfinite(float(value)) for value in values):
            raise FriendlyError("A soundtrack setting is not a finite number.")
        if self.start_time < 0 or self.trim_start < 0 or self.volume < 0 or self.fade_in < 0 or self.fade_out < 0:
            raise FriendlyError("Soundtrack times and volume cannot be negative.")
        if self.trim_end is not None and self.trim_end <= self.trim_start:
            raise FriendlyError("Soundtrack Trim Out must be later than Trim In.")


@dataclass(slots=True)
class ProjectSettings:
    width: int = 1920
    height: int = 1080
    fps: int = 30
    custom_duration: float | None = None
    model_id: str = MODEL_REFERENCE
    visible_cards: int = 0  # Zero means use the model's native layout.
    field_mapping: dict[str, str] = field(default_factory=dict)
    soundtrack_master_volume: float = 1.0

    def effective_visible_cards(self) -> int:
        if self.visible_cards:
            return max(1, min(8, int(self.visible_cards)))
        return MODEL_DEFAULT_VISIBLE.get(self.model_id, VISIBLE_CARDS)

    def auto_duration(self, card_count: int) -> float:
        if card_count <= 0:
            return 0.0
        visible = self.effective_visible_cards()
        reveal = min(card_count, visible) * REFERENCE_REVEAL_SECONDS
        scroll = max(0, card_count - visible) * REFERENCE_SCROLL_SECONDS
        return reveal + scroll + REFERENCE_END_HOLD_SECONDS + REFERENCE_FADE_SECONDS

    def duration(self, card_count: int) -> float:
        automatic = self.auto_duration(card_count)
        if self.custom_duration is None:
            return automatic
        return max(1.0, float(self.custom_duration))

    def speed_multiplier(self, card_count: int) -> float:
        automatic = self.auto_duration(card_count)
        chosen = self.duration(card_count)
        if automatic <= 0 or chosen <= 0:
            return 1.0
        return automatic / chosen

    def model_time(self, output_time: float, card_count: int) -> float:
        return max(0.0, output_time) * self.speed_multiplier(card_count)

    def seconds_per_card(self, card_count: int) -> float:
        speed = self.speed_multiplier(card_count)
        return REFERENCE_SCROLL_SECONDS / speed if speed > 0 else 0.0


@dataclass(slots=True)
class ImportResult:
    cards: list[CardData]
    warnings: list[str] = field(default_factory=list)
    extracted_asset_directory: str | None = None


@dataclass(slots=True)
class TableImportResult:
    data: SpreadsheetData
    warnings: list[str] = field(default_factory=list)
    extracted_asset_directory: str | None = None


@dataclass(slots=True)
class ProjectDocument:
    data: SpreadsheetData
    settings: ProjectSettings
    audio_tracks: list[AudioTrack] = field(default_factory=list)


HEADER_ALIASES = {
    "badge_primary": {
        "date", "uploaded", "upload date", "uploaded date", "year", "value",
        "badge", "badge value", "badge date / value", "number", "amount", "age",
        "probability", "rank",
    },
    "badge_secondary": {
        "unit", "label", "badge label", "badge label / unit", "small label",
        "type", "metric",
    },
    "title": {"title", "name", "heading", "card title", "item", "subject"},
    "description": {"description", "details", "summary", "text", "caption"},
    "image": {"image", "image path", "image url", "photo", "picture", "thumbnail", "artwork"},
}


def normalize_header(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip().lower().replace("_", " "))


def display_cell(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _unique_headers(values: Sequence[object], width: int | None = None) -> list[str]:
    count = width if width is not None else len(values)
    result: list[str] = []
    used: set[str] = set()
    for index in range(count):
        base = display_cell(values[index]) if index < len(values) else ""
        base = base or f"Column {index + 1}"
        candidate = base
        suffix = 2
        while normalize_header(candidate) in used:
            candidate = f"{base} ({suffix})"
            suffix += 1
        result.append(candidate)
        used.add(normalize_header(candidate))
    return result


def guess_field_mapping(headers: Sequence[str]) -> dict[str, str]:
    mapping: dict[str, str] = {}
    normalized = [normalize_header(value) for value in headers]
    for role, aliases in HEADER_ALIASES.items():
        for index, value in enumerate(normalized):
            if value in aliases:
                mapping[role] = headers[index]
                break
    return mapping


def table_from_matrix(rows: Sequence[Sequence[object]], first_row_is_header: bool = True) -> SpreadsheetData:
    matrix = [list(row) for row in rows]
    matrix = [row for row in matrix if any(display_cell(value) for value in row)]
    if not matrix:
        return SpreadsheetData()
    width = max(len(row) for row in matrix)
    if first_row_is_header:
        headers = _unique_headers(matrix[0], width)
        body = matrix[1:]
    else:
        headers = _unique_headers([], width)
        body = matrix
    return SpreadsheetData(
        headers,
        [[display_cell(row[index]) if index < len(row) else "" for index in range(width)] for row in body],
    )


def parse_clipboard_data(text: str) -> SpreadsheetData:
    if not text.strip():
        return SpreadsheetData()
    sample = text[:4096]
    delimiter = "\t" if "\t" in sample else ","
    rows = list(csv.reader(io.StringIO(text), delimiter=delimiter))
    return table_from_matrix(rows, first_row_is_header=True)


def resolve_cards(data: SpreadsheetData, mapping: dict[str, str]) -> list[CardData]:
    normalized = data.normalized()
    indexes = {header: index for index, header in enumerate(normalized.headers)}

    def value(row: Sequence[str], role: str) -> str:
        header = mapping.get(role, "")
        index = indexes.get(header)
        return row[index].strip() if index is not None and index < len(row) else ""

    return [
        CardData(
            uploaded=value(row, "badge_primary"),
            title=value(row, "title"),
            description=value(row, "description"),
            image=value(row, "image"),
            badge_label=value(row, "badge_secondary"),
        )
        for row in normalized.rows
    ]


# Compatibility helpers for 0.1 integrations and tests.
def _header_mapping(values: Sequence[object]) -> dict[str, int]:
    headers = [display_cell(value) for value in values]
    guessed = guess_field_mapping(headers)
    result: dict[str, int] = {}
    role_to_old = {
        "badge_primary": "uploaded", "title": "title", "description": "description", "image": "image"
    }
    for role, old_name in role_to_old.items():
        if role in guessed:
            result[old_name] = headers.index(guessed[role])
    return result


def looks_like_header(values: Sequence[object]) -> bool:
    mapping = _header_mapping(values)
    return "title" in mapping and len(mapping) >= 2


def cards_from_matrix(rows: Sequence[Sequence[object]]) -> list[CardData]:
    cleaned = [list(row) for row in rows if any(display_cell(value) for value in row)]
    if not cleaned:
        return []
    if looks_like_header(cleaned[0]):
        data = table_from_matrix(cleaned, True)
        mapping = guess_field_mapping(data.headers)
    else:
        data = table_from_matrix(cleaned, False)
        mapping = {
            "badge_primary": data.headers[0] if len(data.headers) > 0 else "",
            "title": data.headers[1] if len(data.headers) > 1 else "",
            "description": data.headers[2] if len(data.headers) > 2 else "",
            "image": data.headers[3] if len(data.headers) > 3 else "",
        }
    return [card for card in resolve_cards(data, mapping) if not card.is_blank()]


def parse_clipboard_table(text: str) -> list[CardData]:
    if not text.strip():
        return []
    sample = text[:4096]
    delimiter = "\t" if "\t" in sample else ","
    return cards_from_matrix(list(csv.reader(io.StringIO(text), delimiter=delimiter)))


def format_duration(seconds: float) -> str:
    seconds = max(0, int(round(seconds)))
    hours, remainder = divmod(seconds, 3600)
    minutes, secs = divmod(remainder, 60)
    if hours:
        return f"{hours:02d}:{minutes:02d}:{secs:02d}"
    return f"{minutes:02d}:{secs:02d}"


def parse_duration(value: str) -> float:
    text = value.strip()
    if not text:
        raise FriendlyError(
            "Enter a video duration.",
            "Use seconds, MM:SS, or HH:MM:SS — for example 90, 01:30, or 00:01:30.",
        )
    if re.fullmatch(r"\d+(?:\.\d+)?", text):
        seconds = float(text)
    else:
        parts = text.split(":")
        if len(parts) not in (2, 3) or any(not part.isdigit() for part in parts):
            raise FriendlyError(f'“{value}” is not a valid duration.', "Use seconds, MM:SS, or HH:MM:SS.")
        numbers = [int(part) for part in parts]
        if len(numbers) == 2:
            minutes, secs = numbers
            seconds = minutes * 60 + secs
        else:
            hours, minutes, secs = numbers
            seconds = hours * 3600 + minutes * 60 + secs
        if (minutes >= 60 and len(numbers) == 3) or secs >= 60:
            raise FriendlyError(f'“{value}” is not a valid clock duration.', "Minutes and seconds must be below 60.")
    if not math.isfinite(seconds) or seconds < 1:
        raise FriendlyError("Video duration must be at least one second.")
    return seconds


def _resolve_possible_asset(value: str, workbook_path: Path) -> str:
    if not value or re.match(r"^https?://", value, flags=re.IGNORECASE):
        return value
    if not re.search(r"\.(?:png|jpe?g|webp|gif|tiff?|bmp)$", value, flags=re.IGNORECASE):
        return value
    candidate = Path(value).expanduser()
    if not candidate.is_absolute():
        candidate = workbook_path.parent / candidate
    return str(candidate.resolve())


def load_xlsx_table(path: str | Path) -> TableImportResult:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover
        raise FriendlyError("Excel support is not installed.", "Run: pip install openpyxl") from exc

    workbook_path = Path(path).expanduser().resolve()
    if not workbook_path.is_file():
        raise FriendlyError(f"The workbook does not exist: {workbook_path}", "Choose an existing .xlsx file.")
    try:
        workbook = load_workbook(workbook_path, data_only=True)
    except Exception as exc:
        raise FriendlyError(
            f"Could not open {workbook_path.name}.",
            "Make sure it is a valid, unencrypted .xlsx workbook.",
            str(exc),
        ) from exc

    worksheet = workbook.active
    raw_matrix = [list(row) for row in worksheet.iter_rows(values_only=True)]
    nonblank_indexes = [i for i, row in enumerate(raw_matrix) if any(display_cell(value) for value in row)]
    if not nonblank_indexes:
        raise FriendlyError(f"{workbook_path.name} is empty.", "Add a header row and any data you want to animate.")
    header_sheet_index = nonblank_indexes[0]
    matrix = raw_matrix[header_sheet_index:]
    data = table_from_matrix(matrix, True).normalized()
    for row in data.rows:
        for index, value in enumerate(row):
            row[index] = _resolve_possible_asset(value, workbook_path)

    warnings: list[str] = []
    extracted_directory: str | None = None
    embedded_images = list(getattr(worksheet, "_images", []))
    if embedded_images:
        asset_dir = Path(tempfile.mkdtemp(prefix="comparison-studio-xlsx-"))
        extracted = 0
        for index, embedded in enumerate(embedded_images, start=1):
            try:
                anchor = embedded.anchor._from
                sheet_row = int(anchor.row)
                column = int(anchor.col)
                data_row = sheet_row - header_sheet_index - 1
                if data_row < 0:
                    continue
                while data_row >= len(data.rows):
                    data.rows.append([""] * len(data.headers))
                while column >= len(data.headers):
                    data.headers.append(f"Column {len(data.headers) + 1}")
                    for existing in data.rows:
                        existing.append("")
                raw = embedded._data()
                extension = ".png"
                try:
                    from PIL import Image

                    with Image.open(io.BytesIO(raw)) as opened:
                        extension = f".{(opened.format or 'png').lower()}"
                except Exception:
                    pass
                output = asset_dir / f"embedded_{index:03d}{extension}"
                output.write_bytes(raw)
                data.rows[data_row][column] = str(output)
                extracted += 1
            except Exception as exc:
                warnings.append(f"One embedded image could not be extracted: {exc}")
        if extracted:
            extracted_directory = str(asset_dir)

    if not data.rows:
        warnings.append("The workbook has headers but no data rows yet.")
    return TableImportResult(data.normalized(), warnings, extracted_directory)


def load_xlsx(path: str | Path) -> ImportResult:
    result = load_xlsx_table(path)
    mapping = guess_field_mapping(result.data.headers)
    cards = resolve_cards(result.data, mapping)
    return ImportResult(cards, result.warnings, result.extracted_asset_directory)


def _cards_to_table(cards: Iterable[CardData]) -> SpreadsheetData:
    rows = [[card.uploaded, card.badge_label, card.title, card.description, card.image] for card in cards]
    return SpreadsheetData(["Value", "Label", "Title", "Description", "Image"], rows)


def save_project_json(
    path: str | Path,
    data: SpreadsheetData | Iterable[CardData],
    settings: ProjectSettings,
    audio_tracks: Iterable[AudioTrack] = (),
) -> None:
    table = data if isinstance(data, SpreadsheetData) else _cards_to_table(data)
    payload = {
        "version": 2,
        "spreadsheet": asdict(table.normalized()),
        "settings": asdict(settings),
        "audio_tracks": [asdict(track) for track in audio_tracks],
    }
    Path(path).write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")


def load_project_document(path: str | Path) -> ProjectDocument:
    target = Path(path)
    try:
        payload = json.loads(target.read_text(encoding="utf-8"))
        if int(payload.get("version", 1)) >= 2 and "spreadsheet" in payload:
            data = SpreadsheetData(**payload.get("spreadsheet", {})).normalized()
            settings = ProjectSettings(**payload.get("settings", {}))
            tracks = [AudioTrack(**entry) for entry in payload.get("audio_tracks", [])]
            return ProjectDocument(data, settings, tracks)

        cards = [CardData(**entry) for entry in payload.get("cards", [])]
        data = _cards_to_table(cards)
        old = payload.get("settings", {})
        settings = ProjectSettings(**{key: value for key, value in old.items() if key in ProjectSettings.__dataclass_fields__})
        settings.field_mapping = {
            "badge_primary": "Value",
            "badge_secondary": "Label",
            "title": "Title",
            "description": "Description",
            "image": "Image",
        }
        return ProjectDocument(data, settings, [])
    except Exception as exc:
        raise FriendlyError(
            f"Could not open {target.name}.",
            "Choose a project file created by Comparison Timeline Studio.",
            str(exc),
        ) from exc


def load_project_json(path: str | Path) -> tuple[list[CardData], ProjectSettings]:
    document = load_project_document(path)
    mapping = document.settings.field_mapping or guess_field_mapping(document.data.headers)
    return resolve_cards(document.data, mapping), document.settings
