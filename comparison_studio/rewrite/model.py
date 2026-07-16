from __future__ import annotations

import csv
import io
import json
import math
import re
from dataclasses import asdict, dataclass, field
from pathlib import Path
from typing import Iterable, Sequence

MODEL_REFERENCE = "reference_detail"
MODEL_ILLUSTRATED = "illustrated_cards"
MODEL_COMPACT = "classic_compact"
MODEL_IDS = (MODEL_REFERENCE, MODEL_ILLUSTRATED, MODEL_COMPACT)

MODEL_LABELS = {
    MODEL_REFERENCE: "Reference Detail",
    MODEL_ILLUSTRATED: "Illustrated Cards",
    MODEL_COMPACT: "Classic Compact",
}

HEADER_ALIASES: dict[str, set[str]] = {
    "value": {
        "value",
        "badge",
        "badge value",
        "date",
        "uploaded",
        "upload date",
        "year",
        "rank",
        "probability",
        "amount",
        "age",
    },
    "label": {
        "label",
        "unit",
        "badge label",
        "badge unit",
        "metric",
        "type",
    },
    "title": {"title", "name", "heading", "card title", "item", "subject"},
    "description": {"description", "details", "summary", "text", "caption", "desc"},
    "image": {
        "image",
        "image path",
        "image url",
        "artwork",
        "photo",
        "picture",
        "thumbnail",
        "icon",
    },
}


class ProjectError(RuntimeError):
    """A user-facing project or import error."""


@dataclass(slots=True)
class Card:
    value: str = ""
    label: str = ""
    title: str = ""
    description: str = ""
    image: str = ""

    def is_blank(self) -> bool:
        return not any(
            value.strip()
            for value in (self.value, self.label, self.title, self.description, self.image)
        )


@dataclass(slots=True)
class AudioSettings:
    path: str = ""
    volume: float = 1.0
    loop: bool = False
    fade_in: float = 0.0
    fade_out: float = 0.0

    def normalized(self) -> "AudioSettings":
        return AudioSettings(
            path=str(self.path or "").strip(),
            volume=max(0.0, min(2.0, float(self.volume))),
            loop=bool(self.loop),
            fade_in=max(0.0, float(self.fade_in)),
            fade_out=max(0.0, float(self.fade_out)),
        )


@dataclass(slots=True)
class Project:
    cards: list[Card] = field(default_factory=lambda: [Card() for _ in range(4)])
    model_id: str = MODEL_REFERENCE
    width: int = 1920
    height: int = 1080
    fps: int = 30
    custom_duration: float | None = None
    badge_bounce: bool = True
    audio: AudioSettings = field(default_factory=AudioSettings)
    project_path: str = ""
    dirty: bool = False

    def normalize(self) -> None:
        if self.model_id not in MODEL_IDS:
            self.model_id = MODEL_REFERENCE
        self.width = max(320, min(7680, int(self.width)))
        self.height = max(180, min(4320, int(self.height)))
        self.fps = max(1, min(120, int(self.fps)))
        if self.custom_duration is not None:
            value = float(self.custom_duration)
            self.custom_duration = value if math.isfinite(value) and value > 0 else None
        self.audio = self.audio.normalized()
        self.cards = [card if isinstance(card, Card) else Card(**dict(card)) for card in self.cards]
        if not self.cards:
            self.cards = [Card() for _ in range(4)]

    def content_cards(self) -> list[Card]:
        """Keep deliberate blank rows but trim unused blank rows from the end."""
        cards = list(self.cards)
        while len(cards) > 1 and cards[-1].is_blank():
            cards.pop()
        return cards


def normalize_header(value: object) -> str:
    text = str(value or "").strip().lower().replace("_", " ")
    return re.sub(r"\s+", " ", text)


def _unique_headers(values: Sequence[object], width: int) -> list[str]:
    headers: list[str] = []
    used: set[str] = set()
    for index in range(width):
        base = str(values[index] if index < len(values) else "").strip() or f"Column {index + 1}"
        candidate = base
        suffix = 2
        while normalize_header(candidate) in used:
            candidate = f"{base} ({suffix})"
            suffix += 1
        headers.append(candidate)
        used.add(normalize_header(candidate))
    return headers


def _guess_mapping(headers: Sequence[str]) -> dict[str, int]:
    normalized = [normalize_header(header) for header in headers]
    result: dict[str, int] = {}
    for role, aliases in HEADER_ALIASES.items():
        for index, value in enumerate(normalized):
            if value in aliases:
                result[role] = index
                break
    return result


def _fallback_mapping(width: int) -> dict[str, int]:
    order = ("value", "label", "title", "description", "image")
    return {role: index for index, role in enumerate(order[:width])}


def cards_from_rows(rows: Sequence[Sequence[object]], first_row_is_header: bool = True) -> list[Card]:
    matrix = [list(row) for row in rows]
    matrix = [row for row in matrix if any(str(value or "").strip() for value in row)]
    if not matrix:
        return []

    width = max(len(row) for row in matrix)
    if first_row_is_header:
        headers = _unique_headers(matrix[0], width)
        body = matrix[1:]
        mapping = _guess_mapping(headers)
        if not mapping:
            mapping = _fallback_mapping(width)
    else:
        body = matrix
        mapping = _fallback_mapping(width)

    def value(row: Sequence[object], role: str) -> str:
        index = mapping.get(role)
        if index is None or index >= len(row):
            return ""
        return str(row[index] or "").strip()

    cards = [
        Card(
            value=value(row, "value"),
            label=value(row, "label"),
            title=value(row, "title"),
            description=value(row, "description"),
            image=value(row, "image"),
        )
        for row in body
    ]
    return cards


def parse_table(text: str) -> list[Card]:
    if not text.strip():
        return []
    sample = text[:8192]
    delimiter = "\t" if "\t" in sample else ","
    try:
        rows = list(csv.reader(io.StringIO(text), delimiter=delimiter))
    except csv.Error as exc:
        raise ProjectError(f"Could not read the pasted table: {exc}") from exc
    return cards_from_rows(rows, first_row_is_header=True)


def import_csv(path: str | Path) -> list[Card]:
    target = Path(path).expanduser()
    try:
        text = target.read_text(encoding="utf-8-sig")
    except OSError as exc:
        raise ProjectError(f"Could not open {target.name}: {exc}") from exc
    return parse_table(text)


def import_xlsx(path: str | Path) -> list[Card]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover
        raise ProjectError("XLSX support is unavailable. Install openpyxl.") from exc

    target = Path(path).expanduser().resolve()
    try:
        workbook = load_workbook(target, data_only=True)
    except Exception as exc:
        raise ProjectError(f"Could not open {target.name}: {exc}") from exc
    worksheet = workbook.active
    rows = [list(row) for row in worksheet.iter_rows(values_only=True)]
    cards = cards_from_rows(rows, first_row_is_header=True)

    for card in cards:
        if not card.image or re.match(r"^https?://", card.image, flags=re.IGNORECASE):
            continue
        candidate = Path(card.image).expanduser()
        if not candidate.is_absolute():
            candidate = target.parent / candidate
        card.image = str(candidate.resolve())
    return cards


def _legacy_cards(payload: dict) -> list[Card]:
    if isinstance(payload.get("cards"), list):
        result: list[Card] = []
        for entry in payload["cards"]:
            if not isinstance(entry, dict):
                continue
            result.append(
                Card(
                    value=str(entry.get("value", entry.get("uploaded", "")) or ""),
                    label=str(entry.get("label", entry.get("badge_label", "")) or ""),
                    title=str(entry.get("title", "") or ""),
                    description=str(entry.get("description", "") or ""),
                    image=str(entry.get("image", "") or ""),
                )
            )
        return result

    spreadsheet = payload.get("spreadsheet", {})
    if not isinstance(spreadsheet, dict):
        return []
    headers = [str(value or "") for value in spreadsheet.get("headers", [])]
    rows = spreadsheet.get("rows", [])
    mapping = _guess_mapping(headers)
    if not mapping:
        mapping = _fallback_mapping(len(headers))

    def cell(row: Sequence[object], role: str) -> str:
        index = mapping.get(role)
        if index is None or index >= len(row):
            return ""
        return str(row[index] or "").strip()

    return [
        Card(
            value=cell(row, "value"),
            label=cell(row, "label"),
            title=cell(row, "title"),
            description=cell(row, "description"),
            image=cell(row, "image"),
        )
        for row in rows
        if isinstance(row, (list, tuple))
    ]


def load_project(path: str | Path) -> Project:
    target = Path(path).expanduser().resolve()
    try:
        payload = json.loads(target.read_text(encoding="utf-8"))
    except Exception as exc:
        raise ProjectError(f"Could not open {target.name}: {exc}") from exc
    if not isinstance(payload, dict):
        raise ProjectError(f"{target.name} is not a CTS project.")

    if int(payload.get("format", 0) or 0) >= 3:
        cards = [Card(**entry) for entry in payload.get("cards", []) if isinstance(entry, dict)]
        audio_payload = payload.get("audio", {})
        project = Project(
            cards=cards,
            model_id=str(payload.get("model_id", MODEL_REFERENCE)),
            width=int(payload.get("width", 1920)),
            height=int(payload.get("height", 1080)),
            fps=int(payload.get("fps", 30)),
            custom_duration=payload.get("custom_duration"),
            badge_bounce=bool(payload.get("badge_bounce", True)),
            audio=AudioSettings(**audio_payload) if isinstance(audio_payload, dict) else AudioSettings(),
            project_path=str(target),
        )
    else:
        settings = payload.get("settings", {}) if isinstance(payload.get("settings"), dict) else {}
        audio_tracks = payload.get("audio_tracks", [])
        first_audio = audio_tracks[0] if isinstance(audio_tracks, list) and audio_tracks else {}
        project = Project(
            cards=_legacy_cards(payload),
            model_id=str(settings.get("model_id", MODEL_REFERENCE)),
            width=int(settings.get("width", 1920)),
            height=int(settings.get("height", 1080)),
            fps=int(settings.get("fps", 30)),
            custom_duration=settings.get("custom_duration"),
            badge_bounce=bool(settings.get("hexagons_bounce", True)),
            audio=AudioSettings(
                path=str(first_audio.get("path", "") if isinstance(first_audio, dict) else ""),
                volume=float(first_audio.get("volume", 1.0) if isinstance(first_audio, dict) else 1.0),
                loop=bool(first_audio.get("loop", False) if isinstance(first_audio, dict) else False),
                fade_in=float(first_audio.get("fade_in", 0.0) if isinstance(first_audio, dict) else 0.0),
                fade_out=float(first_audio.get("fade_out", 0.0) if isinstance(first_audio, dict) else 0.0),
            ),
            project_path=str(target),
        )
    project.normalize()
    project.dirty = False
    return project


def save_project(project: Project, path: str | Path) -> str:
    project.normalize()
    target = Path(path).expanduser().resolve()
    payload = {
        "format": 3,
        "application": "Comparison Timeline Studio",
        "cards": [asdict(card) for card in project.cards],
        "model_id": project.model_id,
        "width": project.width,
        "height": project.height,
        "fps": project.fps,
        "custom_duration": project.custom_duration,
        "badge_bounce": project.badge_bounce,
        "audio": asdict(project.audio.normalized()),
    }
    try:
        target.parent.mkdir(parents=True, exist_ok=True)
        target.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")
    except OSError as exc:
        raise ProjectError(f"Could not save {target.name}: {exc}") from exc
    project.project_path = str(target)
    project.dirty = False
    return str(target)


def cards_to_tsv(cards: Iterable[Card]) -> str:
    output = io.StringIO()
    writer = csv.writer(output, delimiter="\t", lineterminator="\n")
    writer.writerow(("Value", "Label", "Title", "Description", "Image"))
    for card in cards:
        writer.writerow((card.value, card.label, card.title, card.description, card.image))
    return output.getvalue()
